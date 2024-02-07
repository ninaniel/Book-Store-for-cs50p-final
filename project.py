import sys
from openpyxl import load_workbook
from tabulate import tabulate
import random
from datetime import datetime
from fpdf import FPDF


def main():
    start_row = 1 
    end_row = start_row + 100
    page = 1
    checkout_items = []
    total_price = 0.0

    greeting()

    while True:
        main_categories_menu()
        try:
            user_choice = int(input("➡ Your choice: "))

            if not 0 < user_choice < 7:
                raise ValueError
            elif user_choice == 6:
                sys.exit("Exiting...\nHope You'll return soon.🤓")
            else:
                print("⏳ Loading 100 books from total of 300\n⏳")      

            while True:
                books = load_books(start_row, end_row, user_choice, page)
                shopping_menu()

                while True:
                    choice_shopping = input("➡ Enter your choice: ").lower()
                    choices = ["a", "b", "c", "d", "e"]
                    if not choice_shopping in choices:
                        print("Invalid choice. Please enter a valid option.") 
                    else:
                        break

                match choice_shopping:
                    case "a":
                        checkout_items, total_price = buy_book(books, checkout_items, total_price)
                        checkout(checkout_items, total_price)
                    case  "b":
                        start_row, end_row, page = next_page(start_row, end_row, page)
                    case  "c":
                        start_row, end_row, page = previous_page(start_row, end_row, page)
                    case  "d":
                        break
                    case  "e":
                        sys.exit("Exiting the program. Thank you!")             

        except ValueError:
            print("Invalid choice. Please enter a valid option.") 


def load_books(start, end, n, page):
    workbook = load_workbook("books.xlsx")
    sheet = workbook[f"Sheet{n}"]

    head_row = list(sheet.values)[0]
    shop_list = list(sheet.values)[start:end]

    print(tabulate(shop_list, headers=head_row, tablefmt="grid" ),
          f"\n📌 Page - {page}")

    return shop_list
      
def buy_book(books, items, price):
    while True:
        try:
            book_choice = int(input("Enter the number of the book you want to buy: "))
            if not book_choice in range(1,101):
                raise ValueError 
            else:
                break
        except ValueError:
            print("⚠ Please, enter the valid number of the book.")
    
    book_index = book_choice - 1
    book_title = books[book_index][1]
    book_price = float(books[book_index][3].removesuffix("$"))

    chosen_book = f"{book_title} - ${book_price}"
    items.append(chosen_book)
    price += book_price

    print(f"\nYou have added: 📖 {chosen_book}")
    return items, price

def checkout(items, price):
    while True:
        checkout_menu(items, price)
        checkout_choice = input("➡ Enter your choice: ")

        match checkout_choice:
            case '1': # checkout
                display_shopping_bag(items)

                while True:
                    pay_choice = input(f"📌 Total Price: ${price}\n\nConfirm payment? (yes/no): ").lower()                           
                    match pay_choice:
                        case 'yes':
                            create_invoice(items, price)
                            sys.exit("Invoice created ✔. Thank you for shopping!")
                        case "no":
                            sys.exit("Payment canceled.\nExiting...")
                        case _:
                            continue
            case '2': # add item - reload page          
                break                
            case '3': # remove item 
                display_shopping_bag(items)
                items, price = remove_item(items, price)
                if len(items) == 0:  
                    print("Shopping bag is empty.")
                    empty_bag_options()
                    break
            case '4': #cancel - emptying bag              
                print("Chekout canceled. Shopping bag is empty.")
                price = 0
                items = []
                empty_bag_options()
                break            
            case _:
                print("Invalid choice. Please enter a valid option.")

def display_shopping_bag(items):
    print("\n📌 Items:")
    for i, item in enumerate(items, start=1):
        print(f"{i}. {item}")

def remove_item(items, price):
    while True:
        try:
            index = int(input("Choose item to remove: "))
            title = items[index-1]
            book_price = float(title.split(" - $")[1])
                                            
            print(f"'{title}' Removed.")
            items.remove(title)

            price -= book_price
            return items, price
            
        except (ValueError, IndexError):
            print("Invalid choice. Please enter a valid option.")

def next_page(start, end, p):
    start += 100
    end += 100
    p += 1
    if p > 3:
        start = 201  # reloading last page
        end =  start + 100
        p = 3
    return  start,  end, p
    
def previous_page(start,  end, p):
    start -= 100
    end -= 100
    p -= 1
    if p == 0:             
        start = 1  # reloading first page
        end =  start + 100
        p = 1
    return  start,  end, p

def greeting():
    print("📚 Welcome to our Book Shop 📚",
          "\n🧐 Searching for some good books to enjoy yourself?",
            "⬇ Have a look at our collections! ⬇", sep="\n")

def main_categories_menu():
    print(  
            "\n➡ Choose a category:"
            "\n1. Books You Must Read Before You Die",
            "2. Science Fiction & Fantasy Books",
            "3. History Books",
            "4. Memoir / Biography / Autobiography",
            "5. Motivational and Self-Improvement Books",
            "6. ❌ Exit", sep = "\n") 
    
def shopping_menu():
    print("\n⚠ Actions:\n",
          "a) 📕 Choose a book to buy",
          "b) 📜 Next page",
          "c) 📜 Previous page",
          "d) 📚 Go Back to Main Collections",
          "e) ❌ Exit", sep= "\n")
    
def checkout_menu(items, price):
    print(f"\n⚠ Items in the shopping bag: {len(items)}",
          f"⚠ Total Price: ${price}",
          "\n1. ✅ Check-out",
          "2. ➕ Add item",
          "3. ➖ Remove item",
          "4. ❌ Cancel", sep= "\n")
    
def empty_bag_options():
    while True:
        choice = input("🛒 Continue shopping? (yes/no): ").lower()
        
        if choice == 'yes':           
            return
        elif choice == 'no':
            sys.exit("Exiting...")
        else:
            continue


def create_invoice(items, price):
    while True:
        name = input("Please, enter your name: ").title()
        address = input("Please, enter your address: ").capitalize()
        if name and address:
            break
           
    pdf = FPDF()
    pdf.add_page()

    # title
    pdf.set_font("Arial", "B", 24)
    pdf.cell(0, 20, "INVOICE", ln=True, align="C")

    # store name
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Book Store", ln=True, align="C")

    # invoice, order info 
    pdf.ln(10)
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, f"Invoice #: {str(random.randint(100000, 999999))}", ln=False, align="L")
    pdf.cell(0, 10, f"Order #: {str(random.randint(100000, 999999))}", ln=True, align="R")
    pdf.cell(0, 10, f"Invoice Date: {datetime.now().strftime('%Y-%m-%d')}", ln=False, align="L")
    pdf.cell(0, 10, f"Order Date: {datetime.now().strftime('%Y-%m-%d')}", ln=True, align="R")

    # costumer info
    pdf.ln(10)
    pdf.cell(0, 7, "Invoice To", ln=True)
    pdf.cell(0, 7, f"Customer Name: {name}", ln=True)
    pdf.cell(0, 7, f"Address: {address}", ln=True)

    # product table
    pdf.ln(10)
    pdf.set_font("Arial", "B", 12)
    pdf.set_fill_color(0, 128, 0)
    page_width = pdf.w - (pdf.l_margin + pdf.r_margin)
    pdf.cell(0.75*page_width, 10, "Product Description", 1, 0, "C", fill=True)
    pdf.cell(0.25*page_width, 10, "Unit Price", 1, 1, "C", fill=True)

    # product details
    pdf.set_font("Arial", "", 12)
    for item in items:
        pdf.cell(0.75*page_width,10, item.split(" - ")[0], 1, 0, "L")
        pdf.cell(0.25*page_width,10, item.split(" - ")[1], 1, 1, "R")

    # total price
    pdf.ln(10)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"Total Price: {price}", ln=True, align="R")

    pdf.output("invoice.pdf")   


if __name__ == "__main__":
    main()